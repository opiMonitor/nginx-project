# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html


# useful for handling different item types with a single interface
# from itemadapter import ItemAdapter


import json
from openpyxl import Workbook
from itemadapter import ItemAdapter

import logging
import dataset
# from sqlalchemy.dialects.postgresql import JSONB
logger = logging.getLogger(__name__)


# do excel
class HostspiderPipeline:
    def process_item(self, item, spider):
        print(f'\nPIPELINE\npipeline host number: {len(item["hosts"])}')
        print(f'pipeline ip number: {len(item["ips"])}\n')

        for index, ip in enumerate(item['ips']):
            item['ips'][index] = ip.split(':')[0]

        wb = Workbook()
        ws = wb.active
        ws.title = 'zabbix_scrapped_hosts'

        for x in range(1, len(item['ips'])+1):
            ws.cell(column=1, row=x, value=item['hosts'][x-1])
            ws.cell(column=2, row=x, value=item['ips'][x-1])

        wb.save('scrapyard.xlsx')

        return item


# do JSON
class JsonWriterPipeline:

    def open_spider(self, spider):
        self.file = open('items.jl', 'w')

    def close_spider(self, spider):
        self.file.close()

    def process_item(self, item, spider):
        line = json.dumps(ItemAdapter(item).asdict()) + "\n"
        self.file.write(line)
        return item


# do PSQL
class PgPipeline(object):

    def __init__(self, crawler):
        from datetime import datetime
        self.process = False
        self.kw = crawler.settings.get('PG_PIPELINE')
        self.bulksize = self.kw.get('bulksize') if self.kw.get('bulksize') else 1000
        self.primary = self.kw.get('primary') if self.kw.get('primary') else None
        self.primary_type = self.kw.get('col').get(
            self.primary)[1] if self.primary and self.kw.get('col').get(self.primary) else None
        self.auto_datetime = self.kw.get('auto_datetime') if self.kw.get(
            'auto_datetime') is not None else False
        self.buffer = []
        self.now = datetime.now()

    @classmethod
    def from_crawler(cls, crawler):
        return cls(crawler)

    def open_spider(self, spider):
        # jesli dane istnieja - usun
        if self.kw.get('connection') and self.kw.get('table_name') and\
                self.kw.get('col') and len(self.kw.get('col')) > 0:
            import dataset
            self.db = dataset.connect(self.kw.get('connection'))
            # automatically carete table
            self.table = self.db.create_table(self.kw.get(
                'table_name'), primary_id=self.primary, primary_type=self.primary_type)
            # automatically create column
            for col_name, (item_col, col_type) in self.kw.get('col').items():
                self.table.create_column(col_name, col_type)
            # automatically create datetime column
            if self.auto_datetime:
                self.table.create_column('datetime', dataset.types.DateTime)
            # delete rows
            self.table.delete()
            # indexing
            if self.primary is not None and self.primary_type is not None:
                self.table.create_index([self.primary])
            if self.kw.get('indexing'):
                for idx in self.kw.get('indexing'):
                    if self.kw.get('col').get(idx):
                        self.table.create_index([idx])

            self.process = True

    def close_spider(self, spider):
        if self.process and len(self.buffer) > 0:
            self.table.insert_many(self.buffer)
            self.buffer = []

    def process_item(self, item, spider):
        if self.process:
            for x in range(0, len(item['ips'])):
                insert_data = {col_name: item.get(
                    item_col)[x] for col_name, (item_col, col_type) in self.kw.get('col').items()}
                if self.auto_datetime:
                    insert_data['datetime'] = self.now
                self.buffer.append(insert_data)
                if len(self.buffer) > self.bulksize:
                    self.table.insert_many(self.buffer)
                    self.buffer = []
        return item
